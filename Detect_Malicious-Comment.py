import pandas as pd
import matplotlib.pyplot as plt
import torch
from transformers import AutoTokenizer, AutoModelForSequenceClassification, TrainingArguments, Trainer
from sklearn.metrics import precision_recall_fscore_support, accuracy_score
from pandas.io.parsers.readers import csv
import time
import openai
from flask import Flask
from flask import request
import openpyxl
import os.path
from os import path

device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
print("device:", device)

def Re_train(self, directory):
  # 댓글과 레이블은 탭("\t")로 구분되어있으므로
  # sep 파라미터로 "\t" 사용
  df = pd.read_csv("/content/korean-malicious-comments-dataset/Dataset.csv", sep="\t")
  df.head()

  null_idx = df[df.lable.isnull()].index
  df.loc[null_idx, "content"]

  # ---- 전처리 ----
  # lable은 content의 가장 끝 문자열로 설정
  df.loc[null_idx, "lable"] = df.loc[null_idx, "content"].apply(lambda x: x[-1])

  #content는 "\t" 앞부분까지의 문자열로 설정
  df.loc[null_idx, "content"] = df.loc[null_idx, "content"].apply(lambda x: x[:-2])

  # 학습을 위해 lable의 데이터타입을 float -> int
  df = df.astype({"lable":"int"})

  df.info()
  # --------

  # Train set / Test set 구분
  train_data = df.sample(frac=0.8, random_state=42)
  test_data = df.drop(train_data.index)

  # 데이터셋 개수 확인
  print('중복 제거 전 학습 데이터셋 : {}'.format(len(train_data)))
  print('중복 제거 전 검증 데이터셋 : {}'.format(len(test_data)))

  # 중복 데이터 제거
  train_data.drop_duplicates(subset=["content"], inplace=True)
  test_data.drop_duplicates(subset=["content"], inplace=True)

  # 데이터셋 개수 확인
  print('중복 제거 후 학습 데이터셋 : {}'.format(len(train_data)))
  print('중복 제거 후 검증 데이터셋 : {}'.format(len(test_data)))

  # 토크나이저 가져오기
  MODEL_NAME = "beomi/KcELECTRA-base"
  tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)

  # 토크나이징
  tokenized_train_sentences = tokenizer(
      list(train_data["content"]),
      return_tensors="pt",
      max_length=128,
      padding=True,
      truncation=True,
      add_special_tokens=True,
  )

  tokenized_test_sentences = tokenizer(
      list(test_data["content"]),
      return_tensors="pt",
      max_length=128,
      padding=True,
      truncation=True,
      add_special_tokens=True,
  )

  train_label = train_data["lable"].values
  test_label = test_data["lable"].values

  train_dataset = CurseDataset(tokenized_train_sentences, train_label)
  test_dataset = CurseDataset(tokenized_test_sentences, test_label)

  # 모델 불러오기
  model = AutoModelForSequenceClassification.from_pretrained(MODEL_NAME, num_labels=2)
  model.to(device)

  # 학습 파라미터 설정
  training_args = TrainingArguments(
      output_dir='./',                # 학습결과 저장경로
      num_train_epochs=10,            # 학습 epoch 설정
      per_device_train_batch_size=8,  # train batch_size 설정
      per_device_eval_batch_size=64,  # test batch_size 설정
      logging_dir='./logs',           # 학습 log 저장경로
      logging_steps=500,              # 학습 log 기록 단위
      save_total_limit=2,             # 학습결과 저장 최대개수
  )


  # 모델의 학습 컨트롤
  trainer = Trainer(
      model=model,                        # 학습하고자 하는 모델
      args=training_args,                 # 위에서 정의한 Training Arguments
      train_dataset=train_dataset,        # 학습 데이터셋
      eval_dataset=test_dataset,          # 평가 데이터셋
      compute_metrics=compute_metrics,    # 평가지표
  )

  # 학습
  trainer.train()

  return True

def Sentence_predict(sent):
  # 평가모드로 변경
  model.eval()

  # 입력된 문장 토크나이징
  tokenized_sent = tokenizer(
      sent,
      return_tensors = "pt",
      truncation=True,
      add_special_tokens=True,
      max_length=128
  )

  # 모델이 위치한 GPU로 이동
  tokenized_sent.to(device)

  # 예측
  with torch.no_grad():
    outputs = model(
        input_ids=tokenized_sent["input_ids"],
        attention_mask=tokenized_sent["attention_mask"],
        token_type_ids=tokenized_sent["token_type_ids"]
    )

  # 결과
  logits = outputs[0]
  logits = logits.detach().cpu()
  result = logits.argmax(-1)
  if result == 0:
    result = True
  elif result == 1:
    result = False
  return result

# 데이터셋 생성
class CurseDataset(torch.utils.data.Dataset):
    def __init__(self, encodings, labels):
        self.encodings = encodings
        self.labels = labels

    def __getitem__(self, idx):
        item = {key: torch.tensor(val[idx]) for key, val in self.encodings.items()}
        item["labels"] = torch.tensor(
            self.labels[idx]
            # RuntimeError: expected scalar type Long but found Int
            , dtype=torch.long
        )
        return item

    def __len__(self):
        return len(self.labels)

# 학습과정에서 사용할 평가지표를 위한 함수 설정
def compute_metrics(pred):
    labels = pred.label_ids
    preds = pred.predictions.argmax(-1)
    precision, recall, f1, _ = precision_recall_fscore_support(labels, preds, average='binary')
    acc = accuracy_score(labels, preds)
    return {
        'accuracy': acc,            # 정확도
        'f1': f1,                   # F1 스코어
        'precision': precision,     # 정밀도
        'recall': recall            # 재현율
    }

def Is_bad_comment(comment):
    # 학습된 데이터를 토대로 악플인지 먼저 확인하고,
    is_bad_comment = Sentence_predict(comment)
    # 악플인 경우 => 악플로 return
    if is_bad_comment:
      return True
    # 악플이 아닌 경우(선플인 경우)
    # chatGPT를 통해 다시 한 번 확인한 후, 학습 여부를 선택
    query = '다음 문장은 악플이야? "' + comment + '"'
    completion = openai.ChatCompletion.create(
      model="gpt-3.5-turbo",
      messages=[
        {"role": "user", "content": query}
      ]
    )
    answer = completion.choices[0].message.content
    print(query, '-', answer)

    if answer.find("네, ") >= 0 or answer.find("예, ") >= 0 or answer.find("악플입니다") >= 0 or answer.find("악플이다") >= 0 or answer.find("악플로 볼 수 있") >= 0:
        return True

    return False


# 댓글과 레이블은 탭("\t")로 구분되어있으므로
# sep 파라미터로 "\t" 사용
df = pd.read_csv("/content/korean-malicious-comments-dataset/Dataset.csv", sep="\t")
df.head()

null_idx = df[df.lable.isnull()].index
df.loc[null_idx, "content"]

# ---- 전처리 ----
# lable은 content의 가장 끝 문자열로 설정
df.loc[null_idx, "lable"] = df.loc[null_idx, "content"].apply(lambda x: x[-1])

#content는 "\t" 앞부분까지의 문자열로 설정
df.loc[null_idx, "content"] = df.loc[null_idx, "content"].apply(lambda x: x[:-2])

# 학습을 위해 lable의 데이터타입을 float -> int
df = df.astype({"lable":"int"})

df.info()
# --------

# Train set / Test set 구분
train_data = df.sample(frac=0.8, random_state=42)
test_data = df.drop(train_data.index)

# 데이터셋 개수 확인
print('중복 제거 전 학습 데이터셋 : {}'.format(len(train_data)))
print('중복 제거 전 검증 데이터셋 : {}'.format(len(test_data)))

# 중복 데이터 제거
train_data.drop_duplicates(subset=["content"], inplace=True)
test_data.drop_duplicates(subset=["content"], inplace=True)

# 데이터셋 개수 확인
print('중복 제거 후 학습 데이터셋 : {}'.format(len(train_data)))
print('중복 제거 후 검증 데이터셋 : {}'.format(len(test_data)))

# 토크나이저 가져오기
MODEL_NAME = "beomi/KcELECTRA-base"
tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)

# 토크나이징
tokenized_train_sentences = tokenizer(
    list(train_data["content"]),
    return_tensors="pt",
    max_length=128,
    padding=True,
    truncation=True,
    add_special_tokens=True,
)

tokenized_test_sentences = tokenizer(
    list(test_data["content"]),
    return_tensors="pt",
    max_length=128,
    padding=True,
    truncation=True,
    add_special_tokens=True,
)

train_label = train_data["lable"].values
test_label = test_data["lable"].values

train_dataset = CurseDataset(tokenized_train_sentences, train_label)
test_dataset = CurseDataset(tokenized_test_sentences, test_label)

# 모델 불러오기
model = AutoModelForSequenceClassification.from_pretrained(MODEL_NAME, num_labels=2)
model.to(device)

# 학습 파라미터 설정
training_args = TrainingArguments(
    output_dir='./',                # 학습결과 저장경로
    num_train_epochs=10,            # 학습 epoch 설정
    per_device_train_batch_size=8,  # train batch_size 설정
    per_device_eval_batch_size=64,  # test batch_size 설정
    logging_dir='./logs',           # 학습 log 저장경로
    logging_steps=500,              # 학습 log 기록 단위
    save_total_limit=2,             # 학습결과 저장 최대개수
)


# 모델의 학습 컨트롤
trainer = Trainer(
    model=model,                        # 학습하고자 하는 모델
    args=training_args,                 # 위에서 정의한 Training Arguments
    train_dataset=train_dataset,        # 학습 데이터셋
    eval_dataset=test_dataset,          # 평가 데이터셋
    compute_metrics=compute_metrics,    # 평가지표
)

# 학습
trainer.train()

trainer.evaluate(eval_dataset=test_dataset)

app = Flask(__name__)

# openai.api_key = "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"


def Training(comment):
    new_filename = 'content/data.xlsx'
    if not path.exists(new_filename):
      # 엑셀 만들기
      wb = openpyxl.Workbook()
    else :
      # 엑셀 열기
      wb = openpyxl.load_workbook(new_filename)

    # 현 시트 선택
    ws = wb.active

    # 엑셀 저장
    ws.append([comment])
    wb.save(new_filename)
    # csv에서 일정 개수 넘어서면 학습
    if ws['A5'].value != None:
      # 쌓인 데이터 학습
      Re_train(new_filename)
      # csv 삭제 후 재생성
      os.remove(new_filename)
      wb = openpyxl.Workbook()
      wb.save(new_filename)
      return True
    return True


@app.route('/analyze', methods=['GET', 'POST'])
def Analyze():
    comment = request.form.get('comment')
    sender = request.form.get('sender')
    regDate = request.form.get('regDate')
    if Is_bad_comment(comment):
        Training(comment)
        return jsonify(
          comment=comment,
          sender=sender,
          regDate=regDate,
          result="bad_comment"
        )
    return jsonify(
        comment=comment,
        sender=sender,
        regDate=regDate,
        result="normal_comment"
    )

@app.route('/')
def hello_world():
    return 'Hello World!'

if __name__ == '__main__':
    app.run()

# while True:
#   sentence = input("댓글을 입력해주세요 : ")
#   if sentence == "quit":
#     break
#   print(sentence_predict(sentence))
#   print("\n")
