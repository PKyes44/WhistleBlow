from pandas.io.parsers.readers import csv
import time
import openai
from flask import Flask
from flask import request
import openpyxl
import os.path
from os import path
from flask_ngrok import run_with_ngrok

app = Flask(__name__)
run_with_ngrok(app)

openai.api_key = "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"


def Training(comment):
    new_filename = 'content/data.xlsx'
    if not path.exists(new_filename):
      # 엑셀 만들기
      wb = openpyxl.Workbook()
    else :
      # 엑셀 열기
      wb = openpyxl.load_workbook(new_filename)

    # 현 시트 선택
    ws = wb.active

    # 엑셀 저장
    ws.append([comment])
    wb.save(new_filename)
    # csv에서 일정 개수 넘어서면 학습
    if ws['A5'].value != None:
      # 쌓인 데이터 학습
      Re_train(new_filename)
      # csv 삭제 후 재생성
      os.remove(new_filename)
      wb = openpyxl.Workbook()
      wb.save(new_filename)
      return True
    return True

def Is_bad_comment(comment):
    # 학습된 데이터를 토대로 악플인지 먼저 확인하고,
    is_bad_comment = Sentence_predict(comment)
    # 악플인 경우 => 악플로 return
    if is_bad_comment:
      return True
    # 악플이 아닌 경우(선플인 경우)
    # chatGPT를 통해 다시 한 번 확인한 후, 학습 여부를 선택
    query = '다음 문장은 악플이야? "' + comment + '"'
    completion = openai.ChatCompletion.create(
      model="gpt-3.5-turbo",
      messages=[
        {"role": "user", "content": query}
      ]
    )
    answer = completion.choices[0].message.content
    print(query, '-', answer)

    if answer.find("네, ") >= 0 or answer.find("예, ") >= 0 or answer.find("악플입니다") >= 0 or answer.find("악플이다") >= 0 or answer.find("악플로 볼 수 있") >= 0:
        return True

    return False

@app.route('/analyze', methods=['GET', 'POST'])
def Analyze():
    comment = request.form.get('comment')
    sender = request.form.get('sender')
    regDate = request.form.get('regDate')
    if Is_bad_comment(comment):
        Training(comment)
        return jsonify(
          comment=comment,
          sender=sender,
          regDate=regDate,
          result="bad_comment"
        )
    return jsonify(
        comment=comment,
        sender=sender,
        regDate=regDate,
        result="normal_comment"
    )

@app.route('/')
def hello_world():
    return 'Hello World!'

if __name__ == '__main__':
    app.run()


